#!/usr/bin/env python3
"""
Python script to write rental property data from JSON to an Excel template.
Preserves existing formulas in the template.

Usage:
    python write_excel.py <input_json_path> <template_excel_path> <output_excel_path>

Worksheet layout (Rental_Income_Worksheet_Template.xlsx — "Rental income" sheet):
    Property columns: E = property 1, F = property 2, G = property 3, ... N = property 10
    Row 5  : property address
    Row 8  : months in service  (Step 1 result — user input)
    Row 11 : rents received     (A1)
    Row 12 : total expenses     (A2)
    Row 13 : insurance          (A3)
    Row 14 : mortgage interest  (A4)
    Row 15 : taxes              (A5)
    Row 16 : HOA dues           (A6, merged with row 17)
    Row 18 : depreciation       (A7)
    Row 19 : extraordinary exp  (A8)
    Rows 20-22, 25, 32, 34, 36-37 : formula-driven — do NOT overwrite
"""

import sys
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Please run: pip install openpyxl")
    sys.exit(1)


def get_column_letter(property_index: int) -> str:
    """
    Get the Excel column letter for a given property index.
    Property 1 -> E, Property 2 -> F, ..., Property 10 -> N
    """
    # Column E is the 5th column (0-indexed: 4)
    column_index = 4 + (property_index - 1)  # 0-indexed: E=4, F=5, G=6, ...

    # Maximum column is N (14th column, 0-indexed: 13)
    if column_index > 13:
        raise ValueError(
            f"Maximum of 10 properties supported (columns E-N). "
            f"Property {property_index} exceeds limit."
        )

    # Convert to Excel column letter
    result = ""
    idx = column_index
    while idx >= 0:
        result = chr(65 + (idx % 26)) + result
        idx = idx // 26 - 1

    return result


def prompt_continue_low_confidence(properties: list) -> bool:
    """Ask user if they want to continue despite low confidence scores."""
    print("\n⚠️  Warning: Some properties have confidence scores below 0.85:")
    for prop in properties:
        if prop.get("confidence", 1.0) < 0.85:
            addr = prop.get("address", "Unknown")
            conf = prop.get("confidence", 0.0)
            print(f"  - {addr}: {conf:.2f}")

    while True:
        response = input("\nDo you want to continue? (y/n): ").strip().lower()
        if response == "y":
            return True
        elif response == "n":
            return False
        else:
            print("Please enter 'y' or 'n'")


def write_json_to_excel(json_path: str, template_path: str, output_path: str) -> None:
    """
    Read JSON data and write it to an Excel template.

    Args:
        json_path: Path to input JSON file
        template_path: Path to Excel template file
        output_path: Path to save the output Excel file
    """
    # Validate input files exist
    json_file = Path(json_path)
    if not json_file.exists():
        raise FileNotFoundError(f"Input JSON file not found: {json_path}")

    template_file = Path(template_path)
    if not template_file.exists():
        raise FileNotFoundError(f"Template Excel file not found: {template_path}")

    # Load JSON data
    with open(json_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Validate JSON structure
    if "properties" not in data:
        raise ValueError("JSON must contain 'properties' array")
    if "validation" not in data:
        raise ValueError("JSON must contain 'validation' object")

    properties = data["properties"]
    validation = data["validation"]

    # Check validation status
    if not validation.get("passed", False):
        notes = validation.get("notes", "Validation failed")
        print(f"❌ Validation failed: {notes}")
        sys.exit(1)

    # Check confidence scores
    low_confidence_exists = any(prop.get("confidence", 1.0) < 0.85 for prop in properties)
    if low_confidence_exists:
        if not prompt_continue_low_confidence(properties):
            print("Operation cancelled by user.")
            sys.exit(0)

    # Load the Excel workbook (do not modify original template)
    print(f"Loading template: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Row mapping from JSON keys to Excel rows
    # Based on Rental_Income_Worksheet_Template.xlsx "Rental income" sheet
    row_mapping = {
        "address":               5,
        "months_in_service":     8,
        "rents_received":        11,
        "total_expenses":        12,
        "insurance":             13,
        "mortgage_interest":     14,
        "taxes":                 15,
        "hoa_dues":              16,
        "depreciation":          18,
        "extraordinary_expense": 19,
    }

    # Formula rows that must NOT be overwritten
    formula_rows = {20, 21, 22, 25, 32, 34, 36, 37}

    # Write each property to the appropriate column
    print(f"Processing {len(properties)} property(ies)...")
    for prop in properties:
        property_index = prop.get("property_index", 1)

        # Get column letter for this property
        column = get_column_letter(property_index)
        addr_preview = str(prop.get("address", "N/A"))[:40]
        print(f"  Writing Property {property_index} ({addr_preview}...) to column {column}")

        # Write each field to the appropriate row
        for json_key, row in row_mapping.items():
            # Safety guard — skip any formula row
            if row in formula_rows:
                continue

            value = prop.get(json_key)

            # Only write if value is not None/null
            if value is not None:
                cell = ws[f"{column}{row}"]
                cell.value = value

    # Save to output path
    print(f"Saving output to: {output_path}")
    wb.save(output_path)
    print("✅ Successfully wrote data to Excel template!")

    # Summary
    print(f"\nSummary:")
    print(f"  - Properties processed: {len(properties)}")
    print(f"  - Output file: {output_path}")


def main():
    """Main entry point."""
    if len(sys.argv) != 4:
        print("Usage: python write_excel.py <input_json_path> <template_excel_path> <output_excel_path>")
        sys.exit(1)

    json_path = sys.argv[1]
    template_path = sys.argv[2]
    output_path = sys.argv[3]

    try:
        write_json_to_excel(json_path, template_path, output_path)
    except FileNotFoundError as e:
        print(f"❌ File Error: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"❌ Value Error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Unexpected Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
